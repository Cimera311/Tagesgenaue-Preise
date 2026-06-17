import openpyxl

# Lade die Excel-Datei
wb = openpyxl.load_workbook('bitcoin2024.xlsx')
print(f"Sheets: {wb.sheetnames}\n")

# Zeige den ersten Sheet
ws = wb.active
print(f"Aktiver Sheet: {ws.title}")
print(f"Dimensionen: {ws.dimensions}\n")

# Zeige die ersten 15 Zeilen
print("Erste 15 Zeilen:")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    print(row)
    if i >= 15:
        break
